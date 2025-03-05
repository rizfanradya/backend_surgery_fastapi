from sqlalchemy import cast, String
from collections import defaultdict
from utils.celery import celery
from utils.database import get_db_session
from utils.retrieve.status import retrieve_status
from utils.send_email_message import send_email, generate_schedule_failed, generate_schedule_success
from utils.parse_date import parse_date
from utils.calculate_week_name import calculate_week_name
from utils.add_duration import add_duration
from models.schedule_queue import ScheduleQueue
from models.schedule_resource import ScheduleResource
from models.procedure_name import ProcedureName
from models.schedule_results import ScheduleResults
from models.ot_assignment import OtAssignment
from models.unit import Unit
from models.day import Day
from models.week import Week
from models.month import Month
from schemas.generate_daily_schedule import ScheduleResourceSchema
from schemas.schedule_results import ScheduleResultsSchema
import json
import os
import pytz
from io import BytesIO
from openpyxl import load_workbook
from datetime import timedelta, datetime, time
from pandas import DataFrame


@celery.task(bind=True, name="utils.tasks.generate_daily_schedule")
def generate_schedule_task(self, schedule_queue_id: int, resource: str):
    with get_db_session() as session:
        indonesia_tz = pytz.timezone('Asia/Jakarta')

        schedule_queue = session.query(ScheduleQueue).where(
            ScheduleQueue.id == schedule_queue_id).first()
        if not schedule_queue:
            return {"status": "error", "message": "Schedule Queue not found"}

        status_failed = retrieve_status('failed', session)
        try:
            email_notif = schedule_queue.user.email_notification
            status_process = retrieve_status('process', session)
            status_completed = retrieve_status('completed', session)

            schedule_queue.status_id = status_process.id
            session.commit()

            resources = [
                ScheduleResourceSchema(**item) for item in json.loads(resource)]
            resource_map = {r.id: r for r in session.query(ScheduleResource).where(
                ScheduleResource.id.in_([item.id for item in resources])).all()}
            for item in resources:
                if item.id in resource_map:
                    for key, value in item.dict().items():
                        if value is not None:
                            setattr(resource_map[item.id], key, value)

            run_ids = session.query(ScheduleResults.run_id).where(
                ScheduleResults.surgery_date.in_(
                    [schedule_queue.start_date, schedule_queue.end_date])
            ).distinct().all()
            run_id_list = [run_id[0] for run_id in run_ids]
            if run_id_list:
                schedule_queue_ids = session.query(ScheduleResults.schedule_queue_id).where(
                    ScheduleResults.run_id.in_(run_id_list)
                ).distinct().all()
                schedule_queue_id_list = [sq[0] for sq in schedule_queue_ids]
                session.query(ScheduleResults).where(ScheduleResults.run_id.in_(
                    run_id_list)).delete(synchronize_session=False)
                if schedule_queue_id_list:
                    session.query(ScheduleQueue).where(ScheduleQueue.id.in_(
                        schedule_queue_id_list)).delete(synchronize_session=False)

            abs_path = os.path.abspath(__file__)
            base_dir = os.path.dirname(os.path.dirname(abs_path))
            upload_dir = os.path.abspath(os.path.join(
                base_dir, '..', 'uploads', 'daily_schedule'))
            os.makedirs(upload_dir, exist_ok=True)
            file_path = os.path.join(upload_dir, schedule_queue.uploaded_file)
            if not os.path.exists(file_path):
                message = f"File {schedule_queue.uploaded_file} not found, file path {file_path}"
                schedule_queue.status_id = status_failed.id
                schedule_queue.log_usr = message
                session.commit()
                if email_notif:
                    send_email(
                        schedule_queue.user.email,
                        'Generate Daily Schedule Failed',
                        generate_schedule_failed(schedule_queue, message)
                    )
                return {"status": "error", "message": message}
            with open(file_path, "rb") as file:
                excel_data = BytesIO(file.read())
            workbook = load_workbook(excel_data)
            sheet = workbook.active
            actual_headers = [cell.value for cell in sheet[1]]

            expected_headers = [
                "BOOKING DATE", "MRN", "AGE", "GENDER", "DIAGNOSIS", "COMMENT",
                "ANAES_TYPE", "TYPE_OF_OPERATION", "SUB_SPECIALITY_DESC", "SPECIALITY_ID",
                "PROCEDURE_NAME", "DURATION", "BOOKED_BY", "SURGEON1", "PACU_REQUIRED",
            ]
            for idx, (expected, actual) in enumerate(zip(expected_headers, actual_headers), start=1):
                if expected != actual:
                    message = f"Column {idx}: Expected '{expected}', found '{actual}'"
                    schedule_queue.status_id = status_failed.id
                    schedule_queue.log_usr = message
                    session.commit()
                    if email_notif:
                        send_email(
                            schedule_queue.user.email,
                            'Generate Daily Schedule Failed',
                            generate_schedule_failed(schedule_queue, message)
                        )
                    return {"status": "error", "message": message}
            procedure_names = {
                p.name for p in session.query(ProcedureName).all()}
            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True),  # type: ignore
                start=2
            ):
                procedure_name = str(row[10])
                if "-" in procedure_name:
                    procedure_name = procedure_name.split("-", 1)[-1].strip()
                procedure_name = f"PROCEDURE - {procedure_name}"
                if procedure_name not in procedure_names:
                    message = f'{procedure_name} in column PROCEDURE NAME and in row {row_idx} not found in database.'
                    schedule_queue.status_id = status_failed.id
                    schedule_queue.log_usr = message
                    session.commit()
                    if email_notif:
                        send_email(
                            schedule_queue.user.email,
                            'Generate Daily Schedule Failed',
                            generate_schedule_failed(schedule_queue, message)
                        )
                    return {"status": "error", "message": message}

            start_date_dt = parse_date(schedule_queue.start_date)
            end_date_dt = parse_date(schedule_queue.end_date)
            operation_dates = []
            current_date = start_date_dt
            while current_date <= end_date_dt:
                if current_date.weekday() < 5:
                    operation_dates.append(current_date)
                current_date += timedelta(days=1)

            ot_time_tracking = {}
            work_day_minutes = 8 * 60
            schedule_results = []

            ot_unit_map = defaultdict(set)
            ot_assignments = session.query(OtAssignment).where(
                OtAssignment.mssp_id == schedule_queue.masterplan_id).all()
            for assignment in ot_assignments:
                ot_unit_map[assignment.unit_id].add(assignment.ot_id)
            ot_unit_map = {unit: list(ots)
                           for unit, ots in ot_unit_map.items()}

            master_plan_weeks = list(
                set(assignment.week_id for assignment in ot_assignments))
            master_plan_weeks.sort()
            num_masterplan_weeks = len(master_plan_weeks)

            master_plan_slots = {(assignment.ot_id, assignment.day_id, assignment.week_id)
                                  : assignment.unit_id for assignment in ot_assignments}
            available_ots = {assignment.ot_id for assignment in ot_assignments}

            for row_idx, row in enumerate([row for row in sheet.iter_rows(  # type: ignore
                    min_row=2, values_only=True)], start=2):
                unit_data = session.query(Unit).where(
                    cast(Unit.name, String).ilike(f"%{row[8]}%")).first()
                if unit_data is None:
                    continue

                duration_str = str(row[11])
                if not duration_str.isdigit() or len(duration_str) != 4:
                    message = f"Invalid duration format at row {row_idx}"
                    schedule_queue.status_id = status_failed.id
                    schedule_queue.log_usr = message
                    session.commit()
                    if email_notif:
                        send_email(
                            schedule_queue.user.email,
                            'Generate Daily Schedule Failed',
                            generate_schedule_failed(schedule_queue, message)
                        )
                    return {"status": "error", "message": message}
                duration_hours = int(duration_str[:2])
                duration_minutes = int(duration_str[2:])
                if not (0 <= duration_hours < 24 and 0 <= duration_minutes < 60):
                    message = "Duration out of valid range"
                    schedule_queue.status_id = status_failed.id
                    schedule_queue.log_usr = message
                    session.commit()
                    if email_notif:
                        send_email(
                            schedule_queue.user.email,
                            'Generate Daily Schedule Failed',
                            generate_schedule_failed(schedule_queue, message)
                        )
                    return {"status": "error", "message": message}
                duration = duration_hours * 60 + duration_minutes
                if duration > work_day_minutes:
                    continue

                procedure_name = row[10]
                if isinstance(procedure_name, str) and "-" in procedure_name:
                    procedure_name = procedure_name.split("-", 1)[-1].strip()
                procedure_name = f"PROCEDURE - {procedure_name}"

                for operation_date in operation_dates:
                    day_id = session.query(Day.id).where(
                        Day.name == operation_date.strftime('%A')
                    ).scalar()
                    if not day_id:
                        message = f"Day not found for date {operation_date}"
                        schedule_queue.status_id = status_failed.id
                        schedule_queue.log_usr = message
                        session.commit()
                        if email_notif:
                            send_email(
                                schedule_queue.user.email,
                                'Generate Daily Schedule Failed',
                                generate_schedule_failed(
                                    schedule_queue, message)
                            )
                        return {"status": "error", "message": message}

                    current_week = session.query(Week).where(
                        Week.name == calculate_week_name(operation_date.date())).first()
                    if not current_week:
                        continue
                    mapped_week_id = master_plan_weeks[(
                        current_week.id - 1) % num_masterplan_weeks]

                    matching_week = session.query(Week).where(
                        Week.name == calculate_week_name(operation_date.date())).first()
                    if not matching_week:
                        continue

                    monday_of_week = operation_date - \
                        timedelta(days=operation_date.weekday())
                    matching_month = session.query(Month).where(
                        cast(Month.name, String).ilike(f"%{monday_of_week.strftime('%B')}%")).first()
                    if not matching_month:
                        message = f"Month not found for date {operation_date}"
                        schedule_queue.status_id = status_failed.id
                        schedule_queue.log_usr = message
                        session.commit()
                        if email_notif:
                            send_email(
                                schedule_queue.user.email,
                                'Generate Daily Schedule Failed',
                                generate_schedule_failed(
                                    schedule_queue, message)
                            )
                        return {"status": "error", "message": message}

                    for ot_id in available_ots:
                        if (ot_id, day_id, mapped_week_id) in master_plan_slots and master_plan_slots[(ot_id, day_id, mapped_week_id)] == unit_data.id:
                            key = (ot_id, day_id, operation_date)
                            if key not in ot_time_tracking:
                                ot_time_tracking[key] = datetime.combine(
                                    operation_date,
                                    time(8, 0)
                                )

                            ot_start_datetime = ot_time_tracking[key]
                            ot_end_datetime = ot_start_datetime + timedelta(
                                minutes=duration)
                            if ot_end_datetime.time() > time(16, 0):
                                continue
                            remaining_minutes = work_day_minutes - (
                                ot_start_datetime.hour * 60 + ot_start_datetime.minute - 480)
                            if duration > remaining_minutes:
                                continue
                            ot_time_tracking[key] = ot_end_datetime

                            schedule_result = ScheduleResultsSchema(
                                run_id=schedule_queue.run_id,
                                schedule_queue_id=schedule_queue.id,
                                unit_id=unit_data.id,
                                mrn=str(row[1]),
                                age=row[2],  # type: ignore
                                week_id=matching_week.id,  # type: ignore
                                mssp_week_id=mapped_week_id,
                                day_id=day_id,
                                month_id=matching_month.id,  # type: ignore
                                surgery_date=ot_start_datetime.date(),
                                type_of_surgery=str(row[7]),
                                sub_specialty_desc=str(row[8]),
                                specialty_id=str(row[9]),
                                procedure_name=procedure_name,
                                surgery_duration=duration,
                                phu_id=unit_data.id,  # type: ignore
                                phu_start_time=(
                                    ot_start_datetime - timedelta(minutes=60)).time(),
                                phu_end_time=ot_start_datetime.time(),
                                ot_id=ot_id,
                                ot_start_time=ot_start_datetime.time(),
                                ot_end_time=ot_end_datetime.time(),
                                surgeon_name=str(row[13]),
                                booked_by=str(row[12]),
                                post_op_id=unit_data.id,  # type: ignore
                                post_op_start_time=ot_end_datetime.time(),
                                post_op_end_time=add_duration(
                                    ot_end_datetime.strftime("%H:%M"), 30, indonesia_tz),
                                pacu_id=unit_data.id,  # type: ignore
                                pacu_start_time=ot_end_datetime.time(),
                                pacu_end_time=add_duration(
                                    ot_end_datetime.strftime("%H:%M"), 90, indonesia_tz),
                                icu_id=unit_data.id,  # type: ignore
                                icu_start_time=add_duration(
                                    ot_end_datetime.strftime("%H:%M"), 90, indonesia_tz),
                                icu_end_time=add_duration(
                                    ot_end_datetime.strftime("%H:%M"), 330, indonesia_tz)
                            )
                            schedule_results.append(
                                ScheduleResults(**schedule_result.dict()))
                            break
                    else:
                        continue
                    break

            schedule_queue.id
            message = "Schedule generated successfully"
            session.add_all(schedule_results)
            schedule_queue.status_id = status_completed.id
            schedule_queue.log_usr = message
            session.commit()
            if email_notif:
                send_email(
                    schedule_queue.user.email,
                    f' - {message}',
                    generate_schedule_success(schedule_queue)
                )

            BASE_DIR = os.path.dirname(os.path.dirname(
                os.path.dirname(os.path.abspath(__file__))))
            DOWNLOAD_DIR = os.path.join(
                BASE_DIR, "downloads", "daily_schedule")
            os.makedirs(DOWNLOAD_DIR, exist_ok=True)
            FILENAME = f"{schedule_queue.run_id}.xlsx"
            FILE_PATH = os.path.join(DOWNLOAD_DIR, FILENAME)
            schedule_result = session.query(ScheduleResults).where(
                ScheduleResults.run_id == schedule_queue.run_id).all()
            data = [{key: value for key, value in r.__dict__.items(
            ) if key != "_sa_instance_state"} for r in schedule_result]
            df = DataFrame(data)
            df.to_excel(FILE_PATH, index=False, engine='openpyxl')

            schedule_queue.downloaded_file = FILENAME
            session.commit()
            return {"status": "success", "message": message}
        except Exception as e:
            schedule_queue.id
            schedule_queue.status_id = status_failed.id
            schedule_queue.log_sys = str(e)
            schedule_queue.log_usr = 'Generate Daily Schedule Failed'
            if email_notif:
                send_email(
                    schedule_queue.user.email,
                    'Generate Daily Schedule Failed',
                    generate_schedule_failed(schedule_queue, '-')
                )
            session.commit()
            return {"status": "error", "message": str(e)}
