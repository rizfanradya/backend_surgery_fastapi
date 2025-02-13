from sqlalchemy import cast, String, func
from utils.celery import celery
from utils.database import get_db_session
from utils.parse_date import parse_date
from utils.calculate_week_name import calculate_week_name
from utils.retrieve.status import retrieve_status
from utils.retrieve.objectives import retrieve_objectives
from utils.retrieve.fixed_ot_type import retrieve_fixed_ot_type
from utils.send_email_message import send_email, generate_masterplan_failed, generate_masterplan_success
from models.masterplan import Masterplan
from models.procedure_name import ProcedureName
from models.clashing_subspecialties import ClashingSubSpecialties
from models.week import Week
from models.ot import Ot
from models.unit import Unit
from models.day import Day
from models.blocked_day import BlockedDay
from models.sub_specialties_ot_types import SubSpecialtiesOtTypes
from models.fixed_ot import FixedOt
from models.blocked_ot import BlockedOt
from models.objectives import Objectives
from models.surgery import Surgery
from models.ot_assignment import OtAssignment
from schemas.surgery import SurgerySchema
from schemas.ot_assignment import OtAssignmentSchema
import os
from io import BytesIO
from openpyxl import load_workbook
from datetime import timedelta, datetime


@celery.task(bind=True, name="utils.tasks.generate_masterplan")
def generate_masterplan_task(self, masterplan_id: int):
    with get_db_session() as session:
        masterplan = session.query(Masterplan).where(
            Masterplan.id == masterplan_id).first()
        if not masterplan:
            return {"status": "error", "message": "Masterplan not found"}

        status_failed = retrieve_status('failed', session)
        try:
            email_notif = masterplan.user.email_notification
            objectives = retrieve_objectives('clashing', session)
            fixed_ot_type = retrieve_fixed_ot_type('fixed', session)
            status_available = retrieve_status('available', session)
            status_process = retrieve_status('process', session)
            status_completed = retrieve_status('completed', session)
            # general_ot_type = session.query(OtType).where(
            #     cast(OtType.description, String).ilike('%gen%')).first()
            # if general_ot_type is None:
            #     send_error_response(
            #         'General ot type not found in database.'
            #     )

            # ult_cln_ot_type = session.query(OtType).where(
            #     cast(OtType.description, String).ilike('%ult%')).first()
            # if ult_cln_ot_type is None:
            #     send_error_response(
            #         'Ultra Clean ot type not found in database.'
            #     )

            abs_path = os.path.abspath(__file__)
            base_dir = os.path.dirname(os.path.dirname(abs_path))
            upload_dir = os.path.abspath(
                os.path.join(base_dir, '..', 'uploads'))
            os.makedirs(upload_dir, exist_ok=True)
            file_path = os.path.join(upload_dir, masterplan.uploaded_file)
            if not os.path.exists(file_path):
                message = f"File {masterplan.uploaded_file} not found, file path {file_path}"
                masterplan.status_id = status_failed.id
                masterplan.log_usr = message
                session.commit()
                if email_notif:
                    send_email(
                        masterplan.user.email,
                        'Generate Masterplan Failed',
                        generate_masterplan_failed(masterplan, message)
                    )
                return {"status": "error", "message": message}
            with open(file_path, "rb") as file:
                excel_data = BytesIO(file.read())
            workbook = load_workbook(excel_data)
            sheet = workbook.active
            actual_headers = [cell.value for cell in sheet[1]]

            procedure_names_db = session.query(ProcedureName).all()
            procedure_names = {p.name for p in procedure_names_db}
            procedure_name_map = {p.name: p.id for p in procedure_names_db}
            ot_names = {o.name for o in session.query(Ot).all()}

            expected_headers = [
                "BOOKING DATE", "OPERATION DATE", "MRN", "AGE", "GENDER CODE", "DIAGNOSIS", "COMMENT",
                "ANAES TYPE", "ANAES ID", "TYPE OF OPERATION", "SUBSPECIALITY DESC", "SPECIALITY ID",
                "OT LIST NAME", "PROCEDURE NAME", "DURATION", "BOOKED BY", "SURGEON"
            ]
            for idx, (expected, actual) in enumerate(zip(expected_headers, actual_headers), start=1):
                if expected != actual:
                    message = f"Column {idx}: Expected '{expected}', found '{actual}'"
                    masterplan.status_id = status_failed.id
                    masterplan.log_usr = message
                    if email_notif:
                        send_email(
                            masterplan.user.email,
                            'Generate Masterplan Failed',
                            generate_masterplan_failed(masterplan, message)
                        )
                    session.commit()
                    return {"status": "error", "message": message}

            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True),  # type: ignore
                start=2
            ):
                procedure_name = str(row[13])
                ot_list_name = str(row[12])
                if "-" in procedure_name:
                    procedure_name = procedure_name.split("-", 1)[-1].strip()
                procedure_name = f"PROCEDURE - {procedure_name}"

                if procedure_name not in procedure_names:
                    message = f'{procedure_name} in column PROCEDURE NAME and in row {row_idx} not found in database.'
                    masterplan.status_id = status_failed.id
                    masterplan.log_usr = message
                    if email_notif:
                        send_email(
                            masterplan.user.email,
                            'Generate Masterplan Failed',
                            generate_masterplan_failed(masterplan, message)
                        )
                    session.commit()
                    return {"status": "error", "message": message}
                if ot_list_name not in ot_names:
                    message = f'{ot_list_name} in column OT LIST NAME and in row {row_idx} not found in database.'
                    masterplan.status_id = status_failed.id
                    masterplan.log_usr = message
                    if email_notif:
                        send_email(
                            masterplan.user.email,
                            'Generate Masterplan Failed',
                            generate_masterplan_failed(masterplan, message)
                        )
                    session.commit()
                    return {"status": "error", "message": message}

            total_weight, num_objectives = session.query(
                func.sum(Objectives.weight),
                func.count(Objectives.id)
            ).one()
            average_weight_obj = total_weight / num_objectives if num_objectives > 0 else 0

            masterplan.objective_value = average_weight_obj
            masterplan.status_id = status_process.id
            session.commit()

            clashing_subspecialties = session.query(
                ClashingSubSpecialties).all()
            clashing_group_map = {}
            for cs in clashing_subspecialties:
                if cs.sub_specialty_id not in clashing_group_map:
                    clashing_group_map[cs.sub_specialty_id] = []
                clashing_group_map[cs.sub_specialty_id].append(cs.unit_id)

            start_date_dt = parse_date(masterplan.start_date)
            end_date_dt = parse_date(masterplan.end_date)
            week_numbers = set()
            current_date = start_date_dt
            while current_date <= end_date_dt:
                week_number = ((current_date - start_date_dt).days // 7) + 1
                week_numbers.add(week_number)
                current_date += timedelta(days=1)
            available_week_ids = []
            for week_number in sorted(week_numbers):
                week_name = f"Week {week_number}"
                week = session.query(Week).where(
                    Week.name == week_name, Week.status_id == status_available.id).first()
                if week:
                    available_week_ids.append(week.id)
            if not available_week_ids:
                message = 'No available weeks found for the specified dates.'
                masterplan.status_id = status_failed.id
                masterplan.log_usr = message
                if email_notif:
                    send_email(
                        masterplan.user.email,
                        'Generate Masterplan Failed',
                        generate_masterplan_failed(masterplan, message)
                    )
                session.commit()
                return {"status": "error", "message": message}
            available_week_ids = session.scalars(
                session.query(Week.id).where(Week.id.in_(
                    available_week_ids))
            ).all()

            surgeries = []
            ot_assignments = []
            ot_assignments_2 = []
            surgeries_data = []
            total_duration_by_unit = {}
            total_days = (end_date_dt - start_date_dt).days + 1
            total_slots = total_days * session.query(Ot).count()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                duration = row[14]
                booking_date = parse_date(row[0])

                unit_data = session.query(Unit).where(
                    cast(Unit.name, String).ilike(f"%{row[10]}%")).first()
                if unit_data is None:
                    continue
                surgeries_data.append({
                    'unit_id': unit_data.id,
                    'duration': duration,
                    'booking_date': booking_date
                })
                if unit_data.id not in total_duration_by_unit:
                    total_duration_by_unit[unit_data.id] = 0
                total_duration_by_unit[unit_data.id] += duration

            total_duration_all = sum(total_duration_by_unit.values())
            slots_allocation = {}
            residual_slots = 0
            for unit_id, total_duration in total_duration_by_unit.items():
                percent = total_duration / total_duration_all if total_duration_all > 0 else 0
                allocated_slots = int(percent * total_slots)
                slots_allocation[unit_id] = allocated_slots
                residual_slots += allocated_slots
            residual_slots = total_slots - sum(slots_allocation.values())
            waiting_times = {}
            for surgery in surgeries_data:
                unit_id = surgery['unit_id']
                waiting_time = (surgery['booking_date'] - start_date_dt).days
                if unit_id not in waiting_times:
                    waiting_times[unit_id] = []
                waiting_times[unit_id].append(waiting_time)
            average_waiting_times = {unit_id: sum(
                times) / len(times) for unit_id, times in waiting_times.items()}
            while residual_slots > 0:
                sorted_units = sorted(
                    average_waiting_times.items(), key=lambda x: x[1], reverse=True)
                for unit_id, _ in sorted_units:
                    if residual_slots > 0:
                        slots_allocation[unit_id] += 1
                        residual_slots -= 1
                    else:
                        break

            for row in sheet.iter_rows(min_row=2, values_only=True):
                unit_data = session.query(Unit).where(
                    cast(Unit.name, String).ilike(f"%{row[10]}%")).first()
                if unit_data is None:
                    continue

                if slots_allocation.get(unit_data.id, 0) > 0:
                    available_ot_ids = []
                    available_day_ids = [day[0] for day in session.query(Day.id).where(Day.id.notin_(
                        {day.day_id for day in session.query(BlockedDay).where(BlockedDay.unit_id == unit_data.id).all()})).all()]

                    unit_fot = session.query(SubSpecialtiesOtTypes).where(
                        SubSpecialtiesOtTypes.unit_id == unit_data.id,
                        SubSpecialtiesOtTypes.sub_specialty_id == unit_data.sub_specialty_id,
                        SubSpecialtiesOtTypes.ot_type_id == fixed_ot_type.id  # type: ignore
                    ).first()

                    if unit_fot is not None:
                        # fixed_ot_query = session.query(FixedOt.ot_id).where(
                        #     FixedOt.unit_id == unit_data.id
                        # )
                        # fixed_ot_type_query = session.query(Ot.id).where(
                        #     Ot.ot_type_id == fixed_ot_type.id  # type: ignore
                        # )
                        # available_ot_ids = [ot[0] for ot in fixed_ot_query.union(
                        #     fixed_ot_type_query).all()]
                        available_ot_ids = [
                            ot[0] for ot in
                            session.query(FixedOt.ot_id).where(
                                FixedOt.unit_id == unit_data.id).all()
                        ]

                    else:
                        # unit_gen_ot = session.query(SubSpecialtiesOtTypes).where(
                        #     SubSpecialtiesOtTypes.unit_id == unit_data.id,
                        #     SubSpecialtiesOtTypes.sub_specialty_id == unit_data.sub_specialty_id,
                        #     SubSpecialtiesOtTypes.ot_type_id == general_ot_type.id  # type: ignore
                        # ).first()
                        # unit_ult_cln_ot = session.query(SubSpecialtiesOtTypes).where(
                        #     SubSpecialtiesOtTypes.unit_id == unit_data.id,
                        #     SubSpecialtiesOtTypes.sub_specialty_id == unit_data.sub_specialty_id,
                        #     SubSpecialtiesOtTypes.ot_type_id == ult_cln_ot_type.id  # type: ignore
                        # ).first()

                        # if unit_gen_ot and unit_ult_cln_ot:
                        #     available_ot_ids = [ot[0] for ot in session.query(Ot.id).where(
                        #         Ot.ot_type_id == general_ot_type.id  # type: ignore
                        #     ).intersect(
                        #         session.query(Ot.id).where(
                        #             Ot.ot_type_id == ult_cln_ot_type.id  # type: ignore
                        #         )
                        #     ).where(
                        #         Ot.id.notin_(
                        #             {bot.ot_id for bot in session.query(BlockedOt)
                        #                 .where(BlockedOt.unit_id == unit_data.id).all()}
                        #         )
                        #     ).all()]
                        # elif unit_gen_ot or unit_ult_cln_ot:
                        #     ot_type_id = general_ot_type.id if unit_gen_ot else ult_cln_ot_type.id  # type: ignore
                        #     available_ot_ids = [ot[0] for ot in session.query(Ot.id).where(
                        #         Ot.ot_type_id == ot_type_id,
                        #         Ot.id.notin_(
                        #             {bot.ot_id for bot in session.query(BlockedOt)
                        #                 .where(BlockedOt.unit_id == unit_data.id).all()}
                        #         )
                        #     ).all()]
                        # else:
                        #     available_ot_ids = []

                        available_ot_ids = [
                            ot[0] for ot in session.query(Ot.id).where(Ot.id.notin_(
                                {bot.ot_id for bot in session.query(BlockedOt).where(
                                    BlockedOt.unit_id == unit_data.id).all()}
                            )).all()
                        ]

                    ot_id = None
                    day_id = None
                    week_id = None
                    daily_ot_counts = {
                        (assignment['day_id'], assignment['unit_id'], assignment['week_id']): 0
                        for assignment in ot_assignments_2
                    }
                    assigned_ots = {
                        (
                            assignment['ot_id'], assignment['day_id'], assignment['week_id']
                        ) for assignment in ot_assignments_2
                    }
                    for assignment in ot_assignments_2:
                        daily_ot_counts[
                            (assignment['day_id'], assignment['unit_id'], assignment['week_id'])] += 1
                    for week_ids in available_week_ids:
                        for day_ids in available_day_ids:
                            if daily_ot_counts.get((day_ids, unit_data.id, week_ids),  # type: ignore
                                                   0) < unit_data.max_slot_limit:
                                for ot_ids in available_ot_ids:
                                    if (ot_ids, day_ids, week_ids) not in assigned_ots:
                                        ot_id = ot_ids
                                        day_id = day_ids
                                        week_id = week_ids
                                        break
                                if ot_id and day_id and week_id:  # type: ignore
                                    break
                            if ot_id and day_id and week_id:  # type: ignore
                                break
                        if ot_id and day_id and week_id:  # type: ignore
                            break
                    if not ot_id or not day_id or not week_id:  # type: ignore
                        continue

                    procedure_name = row[13]
                    if isinstance(procedure_name, str) and "-" in procedure_name:
                        procedure_name = procedure_name.split(
                            "-", 1)[-1].strip()
                    procedure_name = f"PROCEDURE - {procedure_name}"
                    procedure_name_id = procedure_name_map.get(
                        procedure_name, 0)  # type: ignore

                    total_clashes = 0
                    for sub_specialty_id, clashing_units in clashing_group_map.items():
                        if unit_data.id in clashing_units:
                            total_clashes += 1
                    if total_clashes > 0:
                        masterplan.objective_value -= (  # type: ignore
                            total_clashes * objectives.weight)  # type: ignore

                    week_start_date = start_date_dt + \
                        timedelta(weeks=(week_id - min(available_week_ids)))
                    target_date = week_start_date + \
                        timedelta(days=(day_id - 1))
                    if not (start_date_dt <= target_date <= end_date_dt):
                        continue

                    matching_week = session.query(Week).where(
                        Week.name == calculate_week_name(target_date.date())).first()
                    if not matching_week:
                        continue

                    surgery_schema = SurgerySchema(
                        mssp_id=masterplan.id,  # type: ignore
                        mrn=str(row[2]),
                        unit_id=unit_data.id,  # type: ignore
                        booking_date=booking_date,
                        estimated_duration=row[14],  # type: ignore
                        procedure_name_id=procedure_name_id,
                        age=row[3],  # type: ignore
                        gender_code='P' if isinstance(
                            row[4], str) and row[4].upper() == 'P' else 'L',
                        surgeon=row[16]  # type: ignore
                    )
                    ot_assignment_schema = OtAssignmentSchema(
                        mssp_id=masterplan.id,  # type: ignore
                        mrn=str(row[2]),
                        week_id=matching_week.id,
                        week_number=week_id,
                        ot_id=ot_id,  # type: ignore
                        unit_id=unit_data.id,  # type: ignore
                        day_id=day_id,
                        date=target_date,
                        is_require_anaes=True if row[8] == 'Y'else False,
                        opening_time=datetime.strptime(
                            '09:00:00.0000', '%H:%M:%S.%f').time(),
                        closing_time=datetime.strptime(
                            '16:00:00.0000', '%H:%M:%S.%f').time()
                    )
                    surgeries.append(Surgery(**surgery_schema.dict()))
                    ot_assignments.append(OtAssignment(
                        **ot_assignment_schema.dict()))
                    ot_assignments_2.append({
                        'week_id': week_id,
                        'day_id': day_id,
                        'ot_id': ot_id,
                        'unit_id': unit_data.id
                    })
                    slots_allocation[unit_data.id] -= 1

            masterplan.id
            message = "Masterplan generated successfully"
            session.add_all(surgeries)
            session.add_all(ot_assignments)
            masterplan.status_id = status_completed.id
            masterplan.log_usr = message
            session.commit()
            if email_notif:
                send_email(
                    masterplan.user.email,
                    f' - {message}',
                    generate_masterplan_success(masterplan)
                )
            return {"status": "success", "message": message}
        except Exception as e:
            masterplan.id
            masterplan.status_id = status_failed.id
            masterplan.log_sys = str(e)
            masterplan.log_usr = 'Masterplan generated failed'
            if email_notif:
                send_email(
                    masterplan.user.email,
                    'Generate Masterplan Failed',
                    generate_masterplan_failed(masterplan, '-')
                )
            session.commit()
            return {"status": "error", "message": str(e)}
