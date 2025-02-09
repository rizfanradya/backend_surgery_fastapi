from collections import defaultdict
from fastapi import APIRouter, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from utils.database import get_db
from utils.auth import TokenAuthorization
from utils.error_response import send_error_response
from utils.parse_date import parse_date
from utils.retrieve.status import retrieve_status
from utils.tasks.generate_daily_schedule import generate_schedule_task
from utils.remove_orphaned_files import check_and_remove_orphaned_files
from datetime import date, datetime, timedelta
from calendar import monthrange
from openpyxl import load_workbook, Workbook
from pandas import DataFrame
from io import BytesIO
import pytz
import os
from sqlalchemy import func, cast, String, extract
from typing import Optional, Literal
from schemas.schedule_queue import ScheduleQueueSchema
from models.masterplan import Masterplan
from models.schedule_results import ScheduleResults
from models.ot import Ot
from models.unit import Unit
from models.procedure_name import ProcedureName
from models.day import Day
from models.schedule_queue import ScheduleQueue

router = APIRouter()
indonesia_tz = pytz.timezone('Asia/Jakarta')


@router.get('/result')
def schedule_results_and_filter(
    type: Literal['daily', 'weekly', 'monthly'] = 'daily',
    surgery_date: Optional[date] = None,
    week_id: Optional[int] = None,
    month_id: Optional[int] = None,
    year: Optional[int] = None,
    ot_id: Optional[int] = None,
    unit_id: Optional[str] = None,
    surgeon_name: Optional[str] = None,
    patient_name: Optional[str] = None,
    session: Session = Depends(get_db),
    token: str = Depends(TokenAuthorization)
):
    try:
        current_today = datetime.now(indonesia_tz)
        min_year, max_year = session.query(
            func.min(extract('year', ScheduleResults.surgery_date)
                     ).label('min_year'),
            func.max(extract('year', ScheduleResults.surgery_date)
                     ).label('max_year')
        ).first()

        all_months = []
        all_weeks = []
        if min_year and max_year:
            all_months.extend({
                "name": f"{datetime(year, month, 1).strftime('%B')} {year}",
                "month_id": month,
                "year": year
            }
                for year in range(int(min_year), int(max_year) + 1)
                for month in range(1, 13)
            )

            for month in range(1, 13):
                for year in range(int(min_year), int(max_year) + 1):
                    start_date = datetime(year, month, 1)
                    _, days_in_month = monthrange(year, month)
                    end_date = datetime(year, month, days_in_month)
                    current_date = start_date
                    while current_date <= end_date:
                        if current_date.weekday() == 0:
                            week_number = (
                                current_date - start_date).days // 7 + 1
                            week_end = current_date + timedelta(days=4)
                            if week_end.month != current_date.month:
                                week_end = current_date + timedelta(days=4)
                            all_weeks.append({
                                "name": f"{current_date.strftime('%d %b')} - {week_end.strftime('%d %b %Y')}",
                                "fmt_name": f"Week {week_number}: {current_date.strftime('%d %b')} - {week_end.strftime('%d %b %Y')}",
                                "week_id": week_number,
                                "month_id": current_date.month,
                                "start_date": current_date,
                                "end_date": week_end
                            })
                        current_date += timedelta(days=1)

        schedule_results = session.query(
            ScheduleResults,
            Unit.name.label('unit_name'),
            Unit.color_hex.label('unit_color')
        ).join(
            Unit, Unit.id == ScheduleResults.unit_id
        )

        if type == 'daily':
            schedule_results = schedule_results.where(
                ScheduleResults.surgery_date == (surgery_date if surgery_date else date.today()))
        if type == 'weekly':
            if week_id and month_id:
                schedule_results = schedule_results.where(
                    ScheduleResults.week_id == week_id, ScheduleResults.month_id == month_id)
            else:
                current_week_id = (
                    current_today - datetime(
                        current_today.year, current_today.month, 1
                    ).replace(tzinfo=indonesia_tz)
                ).days // 7 + 1
                # if current_today.weekday() in [5, 6]:
                #     current_week_id += 1
                schedule_results = schedule_results.where(
                    ScheduleResults.week_id == current_week_id,
                    ScheduleResults.month_id == current_today.month)
        if type == 'monthly':
            if month_id and year:
                schedule_results = schedule_results.where(
                    ScheduleResults.month_id == month_id,
                    extract('year', ScheduleResults.surgery_date) == year)
            else:
                schedule_results = schedule_results.where(
                    ScheduleResults.month_id == current_today.month,
                    extract('year', ScheduleResults.surgery_date) == current_today.year)
        if ot_id:
            schedule_results = schedule_results.where(
                ScheduleResults.ot_id == ot_id)
        if unit_id:
            schedule_results = schedule_results.where(
                ScheduleResults.unit_id == unit_id)
        if patient_name:
            schedule_results = schedule_results.where(
                cast(ScheduleResults.booked_by, String).ilike(f'%{patient_name}%'))
        if surgeon_name:
            schedule_results = schedule_results.where(
                ScheduleResults.surgeon_name == surgeon_name)

        surgeon_name_list = [
            result.surgeon_name for result, *_ in schedule_results.distinct(
                ScheduleResults.surgeon_name).all()
        ]

        total_data = schedule_results.count()
        schedule_results = schedule_results.all()
        ot_data = session.query(Ot).order_by(Ot.id.asc()).all()

        ot_data_count = {}
        for result, _, _ in schedule_results:
            ot_data_count[result.ot_id] = ot_data_count.get(
                result.ot_id, 0) + 1

        for ot in ot_data:
            count = ot_data_count.get(ot.id, 0)
            ot.category = f"OT {ot.id}\n{count} Surgeries"

        week_date_map = {
            (week['week_id'], week['month_id']): week['fmt_name']
            for week in all_weeks
        }
        month_year_map = {
            (month['year'], month['month_id']): month['name']
            for month in all_months
        }
        fmt_week_date_map = {
            (week['week_id'], week['month_id']): week['name']
            for week in all_weeks
        }

        if type == 'daily':
            formatted_results = []
            for result, unit_name, unit_color in schedule_results:
                get_odc = ot_data_count.get(result.ot_id, 0)
                result.category = f"OT {result.ot_id}\n{get_odc} Surgeries"
                result.surgeries = get_odc
                result.task = f"MRN-{result.mrn}"
                result.color = unit_color
                result.unit_name = unit_name
                formatted_results.append(result)
        else:
            schedule_by_week = defaultdict(list)
            surgeries_count = defaultdict(int)
            for result, _, _ in schedule_results:
                key = (result.week_id, result.day_id, result.ot_id)
                surgeries_count[key] += 1
            for result, unit_name, unit_color in schedule_results:
                key = (result.week_id, result.day_id, result.ot_id)
                year_info = result.surgery_date.year
                result.surgeries = surgeries_count[key]
                result.category = f"OT {result.ot_id}\n{
                    result.surgeries} Surgeries"
                result.task = f"MRN-{result.mrn}"
                result.color = unit_color
                result.unit_name = unit_name
                schedule_by_week[(result.week_id, result.month_id, year_info)].append(
                    result)
            formatted_results = [
                {
                    "week": week_date_map.get((week, month), f"Week {week}"),
                    "fmt_week": fmt_week_date_map.get((week, month), f"Week {week}"),
                    "month": month_year_map.get((year_info, month), f"{month} {year_info}"),
                    "data": data
                }
                for (week, month, year_info), data in schedule_by_week.items()
            ]

        return {
            "total": total_data,
            "all_months": all_months,
            "all_weeks": all_weeks,
            "all_days": session.query(Day).order_by(Day.id.asc()).all(),
            "all_ots": session.query(Ot).order_by(Ot.id.asc()).all(),
            "units": session.query(Unit).order_by(Unit.id.asc()).all(),
            "surgeon_name_list": surgeon_name_list,
            "schedule": formatted_results,
        }
    except Exception as error:
        send_error_response(error, 'Failed to get schedule result')


@router.get('/surgery-details/{mrn}')
def surgery_details(mrn: str, session: Session = Depends(get_db), token: str = Depends(TokenAuthorization)):
    query = session.query(ScheduleResults).where(
        ScheduleResults.mrn == mrn).first()
    if query is None:
        send_error_response('MRN not found')
    return query


@router.post('/generate')
def generate_daily_schedule(
    file: UploadFile = File(...),
    master_plan_id: str = Form(...),
    start_date: date = Form(...),
    end_date: date = Form(...),
    resource: str = Form(...),
    session: Session = Depends(get_db),
    token: str = Depends(TokenAuthorization)
):
    if file.content_type not in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]:
        send_error_response('Wrong file type, only accept xlsx')

    start_date_dt = parse_date(start_date)
    end_date_dt = parse_date(end_date)

    if start_date_dt.weekday() != 0:
        send_error_response('Start date must be a Monday')
    if end_date_dt.weekday() != 4:
        send_error_response('End date must be a Friday')
    if start_date_dt > end_date_dt:
        send_error_response('Start date cannot be after end date')

    master_plan = session.query(Masterplan).where(
        Masterplan.id == master_plan_id).first()
    if master_plan is None:
        send_error_response('Master plan not found')

    status_pending = retrieve_status('pending', session)

    run_id = f"RUN-{int(datetime.now(indonesia_tz).timestamp())}"
    new_schedule_queue_schema = ScheduleQueueSchema(
        run_id=run_id,
        start_date=start_date,
        end_date=end_date,
        masterplan_id=master_plan.id
    )
    new_schedule_queue = ScheduleQueue(**new_schedule_queue_schema.dict())
    new_schedule_queue.status_id = status_pending.id
    new_schedule_queue.user_id = token.id
    session.add(new_schedule_queue)
    session.commit()

    abs_path = os.path.abspath(__file__)
    base_dir = os.path.dirname(os.path.dirname(abs_path))
    upload_dir = os.path.join(base_dir, 'uploads', 'daily_schedule')
    os.makedirs(upload_dir, exist_ok=True)
    file_extension = file.filename.split('.')[-1]
    filename = f'{new_schedule_queue.id}.{file_extension}'
    file_path = os.path.join(upload_dir, filename)
    contents = file.file.read()
    new_schedule_queue.uploaded_file = filename
    session.commit()
    try:
        with open(file_path, 'wb') as f:
            f.write(contents)
    except Exception as error:
        new_schedule_queue.log_info = str(error)
        session.commit()
        send_error_response(str(error), 'Failed to save file')

    task = generate_schedule_task.apply_async(
        args=[new_schedule_queue.id, resource])
    new_schedule_queue.task_id = task.id
    session.commit()
    return {
        "status": "pending",
        "message": "Daily Schedule is being processed",
        "run_id": run_id,
        "task_id": task.id
    }


@router.get('/run_id')
def distinct_run_ids(limit: int = 10, offset: int = 0, session: Session = Depends(get_db), token: str = Depends(TokenAuthorization)):
    try:
        status_completed = retrieve_status('completed', session)
        subquery = (
            session.query(
                ScheduleResults.run_id,
                func.min(ScheduleResults.id).label('min_id')
            )
            .group_by(ScheduleResults.run_id)  # type: ignore
            .order_by(func.min(ScheduleResults.id))
            .subquery()
        )

        return {
            "total": session.query(func.count(subquery.c.run_id)).scalar(),
            "data": [
                {"run_id": row.run_id} for row in (
                    session.query(subquery.c.run_id)
                    .order_by(subquery.c.min_id)  # type: ignore
                    .limit(limit)
                    .offset(offset)
                ).all()
            ],
            "schedule_queue": [
                {
                    "id": sq.id,
                    "run_id": sq.run_id,
                    "uploaded_file": sq.uploaded_file,
                    "created_at": sq.created_at.isoformat(),
                    "log_usr": sq.log_usr,
                    "status": {
                        "id": sq.status.id,
                        "description": sq.status.description
                    } if sq.status else None
                }
                for sq in (
                    session.query(ScheduleQueue)
                    .options(joinedload(ScheduleQueue.status))
                    .filter(ScheduleQueue.status_id != status_completed.id)
                    .all()
                )
            ]
        }
    except Exception as error:
        send_error_response(str(error), 'Failed get run ids')


@router.get('/export')
def export_schedule_results(run_id: str, session: Session = Depends(get_db), token: str = Depends(TokenAuthorization)):
    schedule_result = session.query(ScheduleResults).where(
        ScheduleResults.run_id == run_id).all()
    if not schedule_result:
        send_error_response('Run ID not found')
    data = [r.__dict__ for r in schedule_result]
    df = DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    filename = datetime.now(indonesia_tz).strftime(
        f'schedule_results_{run_id}_%Y-%B-%d_%H:%M:%S.xlsx'
    )
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post('/validity')
def check_excell_format(file: UploadFile = File(...), session: Session = Depends(get_db), token: str = Depends(TokenAuthorization)):
    if file.content_type not in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]:
        send_error_response('Wrong file type, only accept xlsx')
    expected_headers = [
        "BOOKING DATE", "MRN", "AGE", "GENDER", "DIAGNOSIS", "COMMENT",
        "ANAES_TYPE", "TYPE_OF_OPERATION", "SUB_SPECIALITY_DESC", "SPECIALITY_ID",
        "PROCEDURE_NAME", "DURATION", "BOOKED_BY", "SURGEON1", "PACU_REQUIRED",
    ]
    contents = file.file.read()
    excel_data = BytesIO(contents)
    workbook = load_workbook(excel_data)
    sheet = workbook.active
    actual_headers = [cell.value for cell in sheet[1]]  # type: ignore
    for idx, (expected, actual) in enumerate(zip(expected_headers, actual_headers), start=1):
        if expected != actual:
            send_error_response(
                f"Column {idx}: Expected '{expected}', found '{actual}'"
            )
    procedure_names = {p.name for p in session.query(ProcedureName).all()}
    # unit_names = {u.name for u in session.query(Unit).all()}
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),  # type: ignore
        start=2
    ):
        procedure_name = str(row[10])
        subspeciality_desc = str(row[8])
        if "-" in procedure_name:
            procedure_name = procedure_name.split("-", 1)[-1].strip()
        procedure_name = f"PROCEDURE - {procedure_name}"

        if procedure_name not in procedure_names:
            send_error_response(
                f'{procedure_name} in column PROCEDURE NAME and in row {row_idx} not found in database.')
        # if subspeciality_desc not in unit_names:
        #     send_error_response(
        #         f'{subspeciality_desc} in column SUBSPECIALITY DESC and in row {row_idx} not found in database.')
    file.file.seek(0)
    return {'message': 'Excel file is valid with correct headers and data matching the database.'}


@router.get('/template')
def download_template(token: str = Depends(TokenAuthorization)):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "template"  # type: ignore
    headers = [
        "BOOKING DATE", "MRN", "AGE", "GENDER", "DIAGNOSIS", "COMMENT",
        "ANAES_TYPE", "TYPE_OF_OPERATION", "SUB_SPECIALITY_DESC", "SPECIALITY_ID",
        "PROCEDURE_NAME", "DURATION", "BOOKED_BY", "SURGEON1", "PACU_REQUIRED",
    ]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)  # type: ignore
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = datetime.now(indonesia_tz).strftime(
        f'dailyschedule_format_%Y-%B-%d_%H:%M:%S.xlsx')
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.delete('/purge_failed_queues')
def purge_failed_queues(session: Session = Depends(get_db), token: str = Depends(TokenAuthorization)):
    try:
        status_failed = retrieve_status('failed', session)
        session.query(ScheduleResults).where(
            ScheduleResults.schedule_queue_id.in_(
                session.query(ScheduleQueue.id).where(
                    ScheduleQueue.status_id == status_failed.id
                )
            )
        ).delete(synchronize_session=False)
        deleted_rows = session.query(ScheduleQueue).where(
            ScheduleQueue.status_id == status_failed.id
        ).delete(synchronize_session=False)
        session.commit()
        check_and_remove_orphaned_files()
        if deleted_rows == 0:
            return {"status": "info", "message": "No failed queues found"}
        return {"status": "success", "message": f"Deleted {deleted_rows} failed queues"}
    except Exception as error:
        send_error_response(str(error), "Failed to purge queues")
