from fastapi import APIRouter, Depends, File, UploadFile, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from utils.database import get_db
from utils.auth import TokenAuthorization
from utils.transform_ot_type import transform_ot_types
from utils.error_response import send_error_response
from utils.parse_date import parse_date
from utils.tasks.generate_masterplan import generate_masterplan_task
from utils.retrieve.status import retrieve_status
from typing import Literal, Optional
from sqlalchemy import asc, desc, text, cast, String
from io import BytesIO
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
import os
import pytz
from schemas.generate_masterplan import UpdateObjectivesWeightsSchema, ConstraintsResponseSchema, InsertConstraintsSchema
from schemas.sub_specialties_ot_types import SubSpecialtiesOtTypesSchema
from schemas.fixed_ot import FixedOtSchema
from schemas.blocked_ot import BlockedOtSchema
from schemas.preferred_ot import PreferredOtSchema
from schemas.blocked_day import BlockedDaySchema
from schemas.equipment_requirement import EquipmentRequirementSchema
from schemas.masterplan import MasterPlanSchema
from schemas.clashing_subspecialties import ClashingSubSpecialtiesSchema
from models.masterplan import Masterplan
from models.procedure_name import ProcedureName
from models.ot_assignment import OtAssignment
from models.ot import Ot
from models.unit import Unit
from models.week import Week
from models.day import Day
from models.objectives import Objectives
from models.sub_specialties_ot_types import SubSpecialtiesOtTypes
from models.ot_type import OtType
from models.equipment_msp import EquipmentMsp
from models.sub_specialty import SubSpecialty
from models.fixed_ot import FixedOt
from models.blocked_ot import BlockedOt
from models.preferred_ot import PreferredOt
from models.blocked_day import BlockedDay
from models.equipment_requirement import EquipmentRequirement
from models.equipment_requirement_status import EquipmentRequirementStatus
from models.equipment import Equipment
from models.clashing_subspecialties import ClashingSubSpecialties
from models.schedule_resource import ScheduleResource

router = APIRouter()
wib_timezone = pytz.timezone("Asia/Jakarta")


@router.get('/list-master')
def masterplan(
    limit: int = 10,
    offset: int = 0,
    order: Optional[Literal['asc', 'desc']] = 'asc',
    sort_by: Optional[Literal[
        'id',
        'created_at',
        'description',
        'objective_value'
    ]] = 'id',
    session: Session = Depends(get_db),
    token: str = Depends(TokenAuthorization)
):
    query = session.query(Masterplan).options(joinedload(Masterplan.status))
    column_map = {
        'id': Masterplan.id,
        'created_at': Masterplan.created_at,
        'description': Masterplan.description,
        'objective_value': Masterplan.objective_value
    }
    if sort_by in column_map:
        sort_column = column_map[sort_by]
        sort_order = asc if order == 'asc' else desc
        query = query.order_by(sort_order(sort_column))  # type: ignore
    return {
        "total": query.count(),
        "schedule_resource": session.query(ScheduleResource).order_by(ScheduleResource.id.asc()).all(),
        "data": [
            {
                "id": masterplan.id,
                "created_at": masterplan.created_at,
                "description": masterplan.description,
                "objective_value": masterplan.objective_value,
                "uploaded_file": masterplan.uploaded_file,
                "log_usr": masterplan.log_usr,
                "status": {
                    "id": masterplan.status_id,
                    "description": masterplan.status.description
                } if masterplan.status else None
            } for masterplan in query.offset(offset).limit(limit).all()
        ]
    }


@router.get('/constraints', response_model=ConstraintsResponseSchema)
def constraints(session: Session = Depends(get_db), token: str = Depends(TokenAuthorization)):
    try:
        all_ots = session.query(Ot).all()
        all_days = session.query(Day).order_by(Day.id).all()  # type: ignore
        all_equipment_msp = session.query(EquipmentMsp).all()
        all_sub_specialty = session.query(SubSpecialty).all()
        objectives = session.query(Objectives).order_by(
            Objectives.id).all()  # type: ignore
        units = session.query(Unit).options(
            joinedload(Unit.fixed_ot),
            joinedload(Unit.blocked_ot),
            joinedload(Unit.preferred_ot),
            joinedload(Unit.blocked_day),
            joinedload(Unit.equipment_requirement),
            joinedload(Unit.clashing_subspecialties),
        ).order_by(Unit.id).all()  # type: ignore

        day_mapping = {day.id: day.name for day in all_days}
        er_msp_mapping = {e_msp.id: e_msp.name for e_msp in all_equipment_msp}
        ssp_mapping = {ssp.id: ssp.description for ssp in all_sub_specialty}

        for unit in units:
            sub_specialty_ot_type = session.query(SubSpecialtiesOtTypes).outerjoin(OtType).where(
                SubSpecialtiesOtTypes.sub_specialty_id == unit.sub_specialty_id).all()

            unit.ot_types = transform_ot_types(sub_specialty_ot_type, session)
            unit.fixed_ots = [
                {'value': fot.ot_id, 'label': str(fot.ot_id)}
                for fot in unit.fixed_ot
            ]
            unit.fixed_ot_opt = [
                {'value': ot.id, 'label': str(ot.id)}
                for ot in all_ots
            ]

            unit.blocked_ots = [
                {'value': bot.ot_id, 'label': str(bot.ot_id)}
                for bot in unit.blocked_ot
            ]
            unit.blocked_ot_opt = [
                {'value': ot.id, 'label': str(ot.id)}
                for ot in all_ots
            ]

            unit.preferred_ots = [
                {'value': pot.ot_id, 'label': str(pot.ot_id)}
                for pot in unit.preferred_ot
            ]
            unit.preferred_ot_opt = [
                {'value': ot.id, 'label': str(ot.id)}
                for ot in all_ots
            ]

            unit.blocked_days = [
                {
                    'value': bday.day_id,
                    'label': str(day_mapping.get(bday.day_id, 'Unknown'))
                }
                for bday in unit.blocked_day
            ]
            unit.blocked_day_opt = [
                {'value': day.id, 'label': day.name}
                for day in all_days
            ]

            unit.equipment_requirements = [
                {
                    'value': er_msp.equipment_id,
                    'label': str(er_msp_mapping.get(er_msp.equipment_id, 'Unknown'))
                }
                for er_msp in unit.equipment_requirement
            ]
            unit.equipment_requirement_opt = [
                {'value': equipment_msp.id, 'label': equipment_msp.name}
                for equipment_msp in all_equipment_msp
            ]

            unit.sub_specialtys = [
                {
                    'value': ssp.sub_specialty_id,
                    'label': str(ssp_mapping.get(ssp.sub_specialty_id, 'Unknown'))
                }
                for ssp in unit.clashing_subspecialties
            ]
            unit.sub_specialty_opt = [
                {'value': ssp.id, 'label': ssp.description}
                for ssp in all_sub_specialty
                if ssp.id != unit.sub_specialty_id  # type: ignore
            ]

        return {
            'data': {
                'constraints': units,
                'objective': objectives
            }
        }
    except Exception as error:
        send_error_response(
            str(error), 'Failed to get data Constraints or Objectives')


@router.post('/objectives')
def update_objectives_weight(objectives_weights: UpdateObjectivesWeightsSchema, session: Session = Depends(get_db), token: str = Depends(TokenAuthorization)):
    for update in objectives_weights.UpdatesObj:
        data = session.query(Objectives).get(update.id)
        if data is not None:
            if update.weight > 100:
                data.weight = 100
            elif update.weight < 0:
                data.weight = 0
            else:
                data.weight = update.weight
        session.commit()
    return objectives_weights


@router.post('/ins-constraints')
def set_constraints(ins_constraints: InsertConstraintsSchema, session: Session = Depends(get_db), token: str = Depends(TokenAuthorization)):
    truncate_tables = [
        'sub_specialties_ot_types',
        'blocked_day',
        'blocked_ot',
        'fixed_ot',
        'preferred_ot',
        'equipment_requirement',
        'clashing_subspecialties',
    ]

    for constraint in ins_constraints.insConstraints:
        if not any(item.value for key, item in constraint.ot_types.items()):
            send_error_response(
                'At least one ot type should be selected for each unit.'
            )

        blocked_ot_ids = {
            blocked_ot.value for blocked_ot in constraint.blocked_ots
        }
        preferred_ot_ids = {
            preferred_ot.value for preferred_ot in constraint.preferred_ots
        }
        intersection = blocked_ot_ids.intersection(preferred_ot_ids)
        if intersection:
            send_error_response(
                f'Blocked OT and Preferred OT cannot have the same IDs: {
                    intersection}'
            )

    fixed_ot_type = session.query(OtType).where(
        cast(OtType.description, String).ilike('%fix%')).first()
    ers_used_by_some = session.query(EquipmentRequirementStatus).where(
        cast(EquipmentRequirementStatus.description, String).ilike('%used by some%')).first()

    if fixed_ot_type is None or ers_used_by_some is None:
        send_error_response(
            'Fixed ot type or equipment requirement status not found in database.'
        )

    for table in truncate_tables:
        session.execute(text(f'ALTER TABLE {table} DISABLE TRIGGER ALL'))
    for table in truncate_tables:
        session.execute(
            text(f'TRUNCATE TABLE {table} RESTART IDENTITY CASCADE'))
    for table in truncate_tables:
        session.execute(text(f'ALTER TABLE {table} ENABLE TRIGGER ALL'))
    session.commit()

    for constraint in ins_constraints.insConstraints:
        unit_data = session.query(Unit).get(constraint.id)
        if unit_data is not None:
            unit_data.max_slot_limit = constraint.max_slot_limit
            unit_data.no_of_slots = constraint.max_slot_limit if constraint.no_of_slots >= constraint.max_slot_limit else constraint.no_of_slots
            session.commit()

            if fixed_ot_type and any(item.value for key, item in constraint.ot_types.items() if item.id == fixed_ot_type.id):
                for key, item in constraint.ot_types.items():
                    if item.id != fixed_ot_type.id:
                        item.value = False

                fixed_ot_ids = set()
                for fixed_ot in constraint.fixed_ots:
                    ot_data = session.query(Ot).get(fixed_ot.value)
                    if ot_data is not None and fixed_ot.value not in fixed_ot_ids:
                        fixed_ot_schema = FixedOtSchema(
                            unit_id=constraint.id,
                            ot_id=fixed_ot.value
                        )
                        new_fixed_ot = FixedOt(**fixed_ot_schema.dict())
                        session.add(new_fixed_ot)
                        session.commit()
                        fixed_ot_ids.add(fixed_ot.value)
            else:
                blocked_ot_ids = set()
                for blocked_ot in constraint.blocked_ots:
                    ot_data = session.query(Ot).get(blocked_ot.value)
                    if ot_data is not None and blocked_ot.value not in blocked_ot_ids:
                        blocked_ot_schema = BlockedOtSchema(
                            unit_id=constraint.id,
                            ot_id=blocked_ot.value
                        )
                        new_blocked_ot = BlockedOt(**blocked_ot_schema.dict())
                        session.add(new_blocked_ot)
                        session.commit()
                        blocked_ot_ids.add(blocked_ot.value)

                preferred_ot_ids = set()
                for preferred_ot in constraint.preferred_ots:
                    ot_data = session.query(Ot).get(preferred_ot.value)
                    if ot_data is not None and preferred_ot.value not in preferred_ot_ids:
                        preferred_ot_schema = PreferredOtSchema(
                            unit_id=constraint.id,
                            ot_id=preferred_ot.value
                        )
                        new_preferred_ot = PreferredOt(
                            **preferred_ot_schema.dict()
                        )
                        session.add(new_preferred_ot)
                        session.commit()
                        preferred_ot_ids.add(preferred_ot.value)

            for key, item in constraint.ot_types.items():
                ot_type_data = session.query(OtType).get(item.id)
                if item.value and ot_type_data is not None:
                    new_sub_specialties_ot_types_schema = SubSpecialtiesOtTypesSchema(
                        sub_specialty_id=unit_data.sub_specialty_id,
                        ot_type_id=item.id,
                        unit_id=unit_data.id
                    )
                    new_sub_specialties_ot_types = SubSpecialtiesOtTypes(
                        **new_sub_specialties_ot_types_schema.dict()
                    )
                    session.add(new_sub_specialties_ot_types)
                    session.commit()

            blocked_day_ids = set()
            for blocked_day in constraint.blocked_days:
                day_data = session.query(Day).get(blocked_day.value)
                if day_data is not None and blocked_day.value not in blocked_day_ids:
                    blocked_day_schema = BlockedDaySchema(
                        unit_id=constraint.id,
                        day_id=blocked_day.value
                    )
                    new_blocked_day = BlockedDay(**blocked_day_schema.dict())
                    session.add(new_blocked_day)
                    session.commit()
                    blocked_day_ids.add(blocked_day.value)

            equipment_ids = set()
            for equipment_requirement in constraint.equipment_requirements:
                equipment_data = session.query(Equipment).get(
                    equipment_requirement.value)
                if equipment_data is not None and equipment_requirement.value not in equipment_ids:
                    equipment_requirement_schema = EquipmentRequirementSchema(
                        unit_id=constraint.id,
                        equipment_id=equipment_requirement.value,
                        equipment_requirement_status_id=ers_used_by_some.id  # type: ignore
                    )
                    new_equipment_requirement = EquipmentRequirement(
                        **equipment_requirement_schema.dict()
                    )
                    session.add(new_equipment_requirement)
                    session.commit()
                    equipment_ids.add(equipment_requirement.value)

            ssp_ids = set()
            for ssp in constraint.sub_specialtys:
                ssp_data = session.query(SubSpecialty).get(ssp.value)
                if ssp_data is not None and ssp.value not in ssp_ids:
                    existing_clashing = session.query(ClashingSubSpecialties).filter_by(
                        unit_id=constraint.id,
                        sub_specialty_id=ssp.value
                    ).first()
                    if not existing_clashing:
                        ssp_schema = ClashingSubSpecialtiesSchema(
                            unit_id=constraint.id,
                            sub_specialty_id=ssp.value,
                        )
                        new_sub_specialty = ClashingSubSpecialties(
                            **ssp_schema.dict())
                        session.add(new_sub_specialty)
                    ssp_ids.add(ssp.value)
                    reverse_units = session.query(Unit).where(
                        Unit.sub_specialty_id == ssp.value).all()
                    for reverse_unit in reverse_units:
                        if reverse_unit:
                            reverse_clashing_exists = session.query(ClashingSubSpecialties).where(
                                ClashingSubSpecialties.unit_id == reverse_unit.id,
                                ClashingSubSpecialties.sub_specialty_id == unit_data.sub_specialty_id
                            ).first()
                            if not reverse_clashing_exists:
                                reverse_ssp_schema = ClashingSubSpecialtiesSchema(
                                    unit_id=reverse_unit.id,  # type: ignore
                                    sub_specialty_id=unit_data.sub_specialty_id,
                                )
                                reverse_clashing_sub_specialty = ClashingSubSpecialties(
                                    **reverse_ssp_schema.dict())
                                session.add(reverse_clashing_sub_specialty)
                                ssp_ids.add(reverse_unit.id)
            session.commit()
    return {'message': 'Constraints updated successfully.'}


@router.post('/generate')
def generate_masterplan(
    start_date: date = Form(...),
    end_date: date = Form(...),
    file: UploadFile = File(...),
    session: Session = Depends(get_db),
    token: str = Depends(TokenAuthorization)
):
    if file.content_type not in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]:
        send_error_response('Wrong file type, only accept xlsx')

    start_date_dt = parse_date(start_date)
    end_date_dt = parse_date(end_date)

    if start_date_dt.weekday() != 0:
        send_error_response('Start date must be a Monday')
    if end_date_dt.weekday() != 4:
        send_error_response('End date must be a Friday')
    if start_date_dt > end_date_dt:
        send_error_response('Start date cannot be after end date')

    status_pending = retrieve_status('pending', session)
    status_failed = retrieve_status('failed', session)

    new_masterplan_schema = MasterPlanSchema(
        objective_value=0,  # type: ignore
        start_date=start_date,
        end_date=end_date
    )
    new_masterplan = Masterplan(**new_masterplan_schema.dict())
    new_masterplan.status_id = status_pending.id
    new_masterplan.user_id = token.id
    session.add(new_masterplan)
    session.commit()

    current_timestamp = datetime.now(
        wib_timezone).strftime('%Y-%m-%d %H:%M:%S')
    mssp_desc = f'MSSP: {new_masterplan.id} generated at {current_timestamp}'
    abs_path = os.path.abspath(__file__)
    base_dir = os.path.dirname(os.path.dirname(abs_path))
    upload_dir = os.path.join(base_dir, 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    file_extension = file.filename.split('.')[-1]  # type: ignore
    filename = f'{new_masterplan.id}.{file_extension}'
    file_path = os.path.join(upload_dir, filename)
    contents = file.file.read()
    new_masterplan.uploaded_file = filename  # type: ignore
    new_masterplan.description = mssp_desc  # type: ignore
    session.commit()
    try:
        with open(file_path, 'wb') as f:
            f.write(contents)
    except Exception as error:
        message = 'Failed to save file'
        new_masterplan.log_sys = str(error)
        new_masterplan.log_usr = message
        new_masterplan.status_id = status_failed.id
        session.commit()
        send_error_response(str(error), message)

    task = generate_masterplan_task.apply_async(args=[new_masterplan.id])
    new_masterplan.task_id = task.id
    session.commit()
    return {
        "status": "pending",
        "message": "Masterplan is being processed",
        "masterplan_id": new_masterplan.id,
        "task_id": task.id
    }


@router.get('/otassignment')
def otassignment(
    mssp_id: int,
    ot_id: Optional[int] = None,
    unit_id: Optional[int] = None,
    week_id: Optional[int] = None,
    type: Literal['weekly', 'all_weeks'] = 'all_weeks',
    session: Session = Depends(get_db),
    token: str = Depends(TokenAuthorization)
):
    try:
        masterplan = session.query(Masterplan).get(mssp_id)
        if masterplan is None:
            return send_error_response('Masterplan not found.')
        ot_assignment = session.query(OtAssignment).join(
            Day).join(Unit).where(OtAssignment.mssp_id == mssp_id)

        if ot_id:
            ot_assignment = ot_assignment.where(
                OtAssignment.ot_id == ot_id)
        if unit_id:
            ot_assignment = ot_assignment.where(
                OtAssignment.unit_id == unit_id)
        if week_id and type == 'weekly':
            ot_assignment = ot_assignment.where(
                OtAssignment.week_number == week_id)

        ot_assignment = ot_assignment.all()
        ot_assignments_map = {}
        all_ot_ids = [
            ot.id for ot in session.query(Ot).order_by(Ot.id.asc()).all()]

        for assignment in ot_assignment:
            week_id = assignment.week_number
            ot_id = assignment.ot_id
            day_name = assignment.day.name.lower()

            if week_id not in ot_assignments_map:
                ot_assignments_map[week_id] = {}
            if ot_id not in ot_assignments_map[week_id]:
                ot_assignments_map[week_id][ot_id] = {}

            ot_assignments_map[week_id][ot_id][day_name] = {
                "unit_name": assignment.unit.name,
                "is_require_anaes": assignment.is_require_anaes,
                "time": f"{assignment.opening_time} - {assignment.closing_time}",
                "color_hex": assignment.unit.color_hex
            }

        week_date_map = {}
        for assignment in session.query(OtAssignment).where(OtAssignment.mssp_id == mssp_id).all():
            week_id = assignment.week_number
            date = assignment.date
            if week_id not in week_date_map:
                week_date_map[week_id] = {"start_date": date, "end_date": date}
            else:
                if date < week_date_map[week_id]["start_date"]:
                    week_date_map[week_id]["start_date"] = date
                if date > week_date_map[week_id]["end_date"]:
                    week_date_map[week_id]["end_date"] = date

        formatted_weeks = []
        for week in session.query(Week).all():
            if week.id not in week_date_map:
                continue

            start_date = week_date_map[week.id]["start_date"]
            end_date = week_date_map[week.id]["end_date"]

            if start_date.month == end_date.month and start_date.year == end_date.year:
                name = f"{start_date.strftime(
                    '%d')} - {end_date.strftime('%d %b %Y')}"
            elif start_date.year == end_date.year:
                name = f"{start_date.strftime(
                    '%d %b')} - {end_date.strftime('%d %b %Y')}"
            else:
                name = f"{start_date.strftime(
                    '%d %b %Y')} - {end_date.strftime('%d %b %Y')}"

            formatted_weeks.append({
                "name": name,
                "fmt_name": f"{week.name}: {name}",
                "week_id": week.id,
                "month_id": start_date.month,
                "start_date": start_date.strftime('%Y-%m-%d'),
                "end_date": end_date.strftime('%Y-%m-%d')
            })
        formatted_weeks = sorted(
            formatted_weeks, key=lambda x: datetime.strptime(x['start_date'], '%Y-%m-%d'))

        grouped_data = []
        for week_id, ot_data in ot_assignments_map.items():
            data = []
            for ot_id in all_ot_ids:
                week = ot_data.get(ot_id, {})
                data.append({
                    "ot_id": ot_id,
                    "week": week
                })
            grouped_data.append({
                "week": next((fw for fw in formatted_weeks if fw['week_id'] == week_id), {}),
                "data": data
            })

        if not grouped_data:
            selected_week = next(
                (fw for fw in formatted_weeks if fw['week_id'] == week_id), {})
            grouped_data.append({
                "week": selected_week,
                "data": []
            })

        return {
            "otassignment": grouped_data,
            "masterPlan": masterplan,
            "weeks": formatted_weeks,
            "days": session.query(Day).order_by(Day.id.asc()).all(),
            "ots": session.query(Ot).order_by(Ot.id.asc()).all(),
        }
    except Exception as error:
        send_error_response(error, 'Cannot get ot assignment data')


@router.post('/validity')
def check_excell_format(file: UploadFile = File(...), session: Session = Depends(get_db), token: str = Depends(TokenAuthorization)):
    if file.content_type not in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]:
        send_error_response('Wrong file type, only accept xlsx')

    expected_headers = [
        "BOOKING DATE", "OPERATION DATE", "MRN", "AGE", "GENDER CODE", "DIAGNOSIS", "COMMENT",
        "ANAES TYPE", "ANAES ID", "TYPE OF OPERATION", "SUBSPECIALITY DESC", "SPECIALITY ID",
        "OT LIST NAME", "PROCEDURE NAME", "DURATION", "BOOKED BY", "SURGEON"
    ]

    contents = file.file.read()
    excel_data = BytesIO(contents)
    workbook = load_workbook(excel_data)
    sheet = workbook.active
    actual_headers = [cell.value for cell in sheet[1]]  # type: ignore

    for idx, (expected, actual) in enumerate(zip(expected_headers, actual_headers), start=1):
        if expected != actual:
            send_error_response(
                f"Column {idx}: Expected '{expected}', found '{actual}'"
            )
    procedure_names = {p.name for p in session.query(ProcedureName).all()}
    ot_names = {o.name for o in session.query(Ot).all()}
    # unit_names = {u.name for u in session.query(Unit).all()}

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),  # type: ignore
        start=2
    ):
        procedure_name = str(row[13])
        ot_list_name = str(row[12])
        # subspeciality_desc = str(row[10])
        # operation_date = parse_date(row[1])

        if "-" in procedure_name:
            procedure_name = procedure_name.split("-", 1)[-1].strip()
        procedure_name = f"PROCEDURE - {procedure_name}"

        if procedure_name not in procedure_names:
            send_error_response(
                f'{procedure_name} in column PROCEDURE NAME and in row {row_idx} not found in database.')

        if ot_list_name not in ot_names:
            send_error_response(
                f'{ot_list_name} in column OT LIST NAME and in row {row_idx} not found in database.')

        # if subspeciality_desc not in unit_names:
        #     send_error_response(
        #         f'{subspeciality_desc} in column SUBSPECIALITY DESC and in row {row_idx} not found in database.')
    file.file.seek(0)
    return {'message': 'Excel file is valid with correct headers and data matching the database.'}


@router.get('/template')
def download_template(token: str = Depends(TokenAuthorization)):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "template"  # type: ignore
    headers = [
        "BOOKING DATE", "OPERATION DATE", "MRN", "AGE", "GENDER CODE", "DIAGNOSIS", "COMMENT",
        "ANAES TYPE", "ANAES ID", "TYPE OF OPERATION", "SUBSPECIALITY DESC", "SPECIALITY ID",
        "OT LIST NAME", "PROCEDURE NAME", "DURATION", "BOOKED BY", "SURGEON"
    ]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)  # type: ignore
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = datetime.now(wib_timezone).strftime(
        f'masterplan_format_%Y-%B-%d_%H:%M:%S.xlsx')
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
